import os
import datetime
import openpyxl
import compaire
import json
import re
import xlsxwriter
from itertools import zip_longest
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.styles.colors import BLUE
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.hyperlink import Hyperlink

from openpyxl import load_workbook
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
import opg
class get_casefile ():
    def __init__(self):
        self.case_dir=r'resource/basic'
        self.case_file_list=[]
        self.txt='resource/test_report/report.txt'
        self.cons_txt='resource/test_report/cons.txt'
        #self.err_txt='resource/test_report/cons_err.txt'
        self.bfe=None
        self.cons=None
        self.cons_err=[]
    def ergodic(self):
        #p=os.listdir(self.case_dir)
        for files in os.listdir(self.case_dir):
            if files.endswith('.sql'):
                self.case_file_list.append('%s/%s'%(self.case_dir,files))
        print(self.case_file_list)
        return self.case_file_list
    def analysis(self,err):

        x = {'P': '', 'cons_diff': []}
        a='YES 源备无差异'

        with open(self.txt,'r+') as f:
            self.bfe=json.load(f)
        with open(self.cons_txt,'r+') as c:
            self.cons=json.load(c)
        # with open(self.cons_txt,'r+') as e:
        #     err=json.load(e)
        for temp in err:
            if len(temp)==0:
                x['P']=a
                self.cons_err.append(x)
            else:
                for i in temp:
                    x['cons_diff'].append(ILLEGAL_CHARACTERS_RE.sub('非法字符',f'{i[0][0]}行 {i[0][1]}列，源：{i[1]} 备：{i[2]}'))
                x['P']=f'NO 源备存在差异，共{len(temp)}个'
                self.cons_err.append(x)
            x = {'P': '', 'cons_diff': []}
        #print(type(self.bfe),self.bfe)
    def more(self):
        pass
    def num_excel(self,number):
        result = ''
        while number > 0:
            # 获取当前数字对应的字母
            remainder = (number - 1) % 26
            result = chr(65 + remainder) + result
            # 更新数字
            number = (number - 1) // 26
        return result
    def xlsx(self,src,db):
        bfe=self.bfe
        cs=self.cons
        err=self.cons_err
        seq=0
        now = datetime.datetime.now().strftime("%Y-%m-%d %H").replace('-', '').replace(' ', '')
        wb=openpyxl.load_workbook(r'resource/test_report/report.xlsx')
        sheet_name=wb.sheetnames
        while now+db+str(seq) in sheet_name:
            seq+=1

        sn=now+db+str(seq)
        seq = 0
        ws = wb.create_sheet(sn)
        ws.cell(1,1).value='%s -- %s'%(src,now+db);ws.cell(1,5).value=bfe[0];ws.cell(2,1).value='表名';ws.cell(2,2).value='pk';ws.cell(2,11).value='状态'
        ws.cell(2, 3).value = 'default';ws.cell(2, 4).value = '唯一';ws.cell(2, 5).value = '复合唯一';ws.cell(2, 6).value = 'check';ws.cell(2, 7).value = '源ck_def'
        ws.cell(2, 8).value = '备check';ws.cell(2, 9).value = '索引';ws.cell(2, 10).value = '复合索引';ws.cell(2, 23).value = '表内容'
        blue=Font(color=Color(rgb=BLUE))
        red= Font(color=Color(rgb="FF0000"))
        l=0
        h=4
        i=j=k=3
        x_int = 3#详情25列第三行 x-int 行 x_y 列
        x_y=25
        #s=len(bfe[1::])-len(cs)
        for item,co in zip_longest(bfe[1::],cs):
            ws.cell(i, 1).value = item[0]
            if '备端查无该表' in item[1::][0]:
                i+=1;h+=1;continue
            #for x in item[1::]:print(x)
            # if not co:
            #     ws.cell(i, 11).value = ILLEGAL_CHARACTERS_RE.sub(r'', cell)
            for cell,col in zip_longest(item[1::],co):#第三列-第n行  字段约束属性
                if not cell: i += 1;h += 1;continue
                    #
                ws.cell(i, 11).value = ILLEGAL_CHARACTERS_RE.sub(r'', cell)
                #print(i,cell)
                if col:
                    for c in range(1,len(col)+1):
                        if not col[c-1]:
                            a='NULL'
                        else:
                            a=col[c-1]
                        ws.cell(h, c+1).value = ILLEGAL_CHARACTERS_RE.sub(r'', a)
                else:pass
                    #print(a)
                i+=1;h+=1

        while j<=i:#按表名合并相同行
            s=ws.cell(k, 1).value
            a=ws.cell(j,1).value
            if ws.cell(j,1).value is None or a==s  :j+=1
            else:
                ws.merge_cells(start_row=k, end_row=j-1, start_column=1, end_column=1)

                if err[l]['P'].startswith('Y'):
                    ws.cell(k , 23).value = err[l]['P']
                    ws.cell(k, 24).value = '=A' + str(k)
                    #ws.cell(k+1, 23).value = err[l]['cons_diff']
                else:
                    col_for=self.num_excel(x_y)+str(x_int-1)#x_int行x_y列
                    ws.cell(k,24).value = '=A' + str(k)
                    col_bak='W'+str(k)
                    hyper_for=f'''=HYPERLINK("#{sn}!{col_for}","{err[l]['P']}")'''
                    hyper_back=f'''=HYPERLINK("#{sn}!{col_bak}","{bfe[1::][l][0]}")'''
                    ws.cell(k , 23).value = hyper_for;ws.cell(k , 23).font=blue#加超链接，有问题项
                    ws.cell(x_int - 1, x_y).value = hyper_back;ws.cell(x_int-1,x_y).font=blue
                    #ws.cell(k , 23).hyperlink = Hyperlink(target=ws.cell(x_int-1, x_y).coordinate)
                    #ws.cell(x_int-1,x_y).hyperlink = Hyperlink(target=ws.cell(row=k, column=23).coordinate,location='返回')
                    for xi in err[l]['cons_diff']:
                        ws.cell(x_int,x_y).value = xi
                        x_int+=1
                    x_y+=1
                    x_int=3
                    #ws.cell(k+1, 23).value = err[l]['cons_diff']
                k=j
                l+=1#问题表index
        wb.save(filename=r'resource/test_report/report.xlsx');print('结果存至',sn,'工作表')
#cell.hyperlink = Hyperlink(target=ws.cell(row=2, column=2).coordinate, location='详情')

#     pass
if __name__ == '__main__':
    #get_casefile().ergodic()
    err=[]
    now = datetime.datetime.now().strftime("%Y-%m-%d %H").replace('-','').replace(' ','')
    print(now)
    p=get_casefile()
    c=compaire.ergodic_database()
    #p.xlsx()
    p.analysis(err)
    p.xlsx('postgre')