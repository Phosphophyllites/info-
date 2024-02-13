import numpy
import datetime
import cx_Oracle
from shapely import wkb
import pandas
from pandas import NaT
from datetime import datetime,timezone,timedelta,date,time
from dateutil.tz import tzoffset
import struct
from openpyxl.worksheet.hyperlink import Hyperlink
import subprocess
import multiprocessing
from threading import Thread

import openpyxl
def tyop(row,src_cube,tgt_cube):

    if src_cube == tgt_cube:
        row.append((1, type(src_cube)))
    elif ((type(src_cube) and type(tgt_cube)) == bytes) and src_cube is tgt_cube:  # byte类型处理
        row.append((1, type(src_cube)))
    # # geometry处理
    # elif isinstance(tgt_cube, cx_Oracle.LOB) and isinstance(src_cube, bytes):
    #     shapely_geometry = wkb.loads(src_cube, hex=True)
    #     # 检查几何类型
    #     if shapely_geometry.geom_type == 'Point':
    #         # 获取点坐标
    #         point_coordinates = shapely_geometry.coords.xy
    #         x, y = point_coordinates[0][0], point_coordinates[1][0]
    #         print(f"点坐标：({x}, {y})")
    # time类型处理
    elif type(tgt_cube) == pandas._libs.tslibs.timestamps.Timestamp and type(
            src_cube) == bytes:  # ssms为datetimeoffset，特殊处理
        unpacked = struct.unpack('QIhH', src_cube)
        m = []
        for tup in unpacked:
            m.append(tup)
        days = m[1]
        microseconds = m[0] / 10 if m[0] else 0
        timezone = m[2]
        tz = tzoffset('ANY', timezone * 60)
        my_date = datetime(*[1900, 1, 1, 0, 0, 0], tzinfo=tz)
        td = timedelta(days=days, minutes=m[2], microseconds=microseconds)
        my_date += td
        print(type(my_date), my_date)
        if str(my_date)[:-6] == str(tgt_cube):
            row.append((1, type(src_cube)))
        else:
            row.append((0, type(src_cube)))
    elif isinstance(tgt_cube, pandas._libs.tslibs.nattype.NaTType) and not src_cube:
        row.append((1, type(src_cube)))
    elif type(src_cube) == pandas._libs.tslibs.timestamps.Timestamp:  # ssms timestamp精度可能丢失，保留小数2位再做比较
        if str(src_cube)[:-4] == str(tgt_cube)[:-4]:
            row.append((1, type(src_cube)))
        else:
            row.append((0, type(src_cube)))
    elif isinstance(src_cube, time):  # ssms time毫秒有误差，只判断至十分位
        if str(src_cube)[:11] == str(tgt_cube)[:11]:
            row.append((1, type(src_cube)))
        else:
            row.append((0, type(src_cube)))
    elif isinstance(src_cube, date):  # ssms date转换至str直接比较
        if str(src_cube) == str(tgt_cube):
            row.append((1, type(src_cube)))
        else:
            row.append((0, type(src_cube)))

    elif type(tgt_cube) == numpy.float64:  # ssms float精度不一致处理
        if str(src_cube)[0:len(str(tgt_cube))] == str(tgt_cube):
            row.append((1, type(src_cube)))
        else:
            row.append((0, type(src_cube)))
    # binary类型
    elif isinstance(tgt_cube, cx_Oracle.LOB):
        if tgt_cube.read() == src_cube:
            row.append((1, type(src_cube)))
        else:#geometry
            shapely_geometry = wkb.loads(src_cube, hex=True)
            # 检查几何类型
            if shapely_geometry.geom_type == 'Point':
                # 获取点坐标
                point_coordinates = shapely_geometry.coords.xy
                x, y = point_coordinates[0][0], point_coordinates[1][0]
                print(f"点坐标：({x}, {y})")

        # print(1,type(src_cube))
    else:
        row.append((0, type(src_cube)))
    return row


    #print(row)
def bin(src,tgt):


    #return row
    pass

if __name__ == '__main__':
    wb = openpyxl.load_workbook(r'resource/test_report/report.xlsx')
    ws = wb.create_sheet('1')
    #wb.save(r'resource/test_report/report.xlsx')

    #ws.cell(1, 23).hyperlink = Hyperlink(target=ws.cell('A1').coordinate)
    wb.save(r'resource/test_report/report.xlsx')